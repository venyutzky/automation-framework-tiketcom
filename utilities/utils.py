import softest
import logging
import inspect
from openpyxl import Workbook, load_workbook
class  Utils(softest.TestCase):
            
    def assertListItemText(self, list, value):
        for item in list:
            print("The Filter Transit is " + item.text)
            self.soft_assert(self.assertEqual, item.text, value)
            if item.text == value:
                print("test passed")
            else:
                print("test failed")
        self.assert_all()
        
    def custom_logger(loglevel=logging.DEBUG):
        #Set class/method name from wehere its called
        logger_name = inspect.stack()[1][3]
        #Create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        #Create console handler or file handler and set the log level
        fh = logging.FileHandler("automation.log", mode="a")
        #Create formatter - how you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        #Add formatter to console or file handler
        fh.setFormatter(formatter)
        #Add console handler to logger
        logger.addHandler(fh)
        return logger
    
    def read_data_from_excel(filename, sheet):
        datalist = []
        wb = load_workbook(filename=filename)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column
        
        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist